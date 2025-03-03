import sys
from PyQt5.QtWidgets import *
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QFont, QPalette, QColor
import os
import shutil
import glob
from PIL import Image
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from PyQt5.QtCore import QDateTime

import os
import glob
import cv2
import numpy as np
from tqdm import tqdm
import ExcelManager
from openpyxl.drawing.image import Image as xlImage
import random
from io import BytesIO

# 新增Excel管理类
class ExcelManager:
    def __init__(self, file_path):
        self.wb = load_workbook(file_path)
        self.ws = self.wb.active
        self.file_path = file_path
        self.current_row = 1
        self.max_width = 1024  # 最大图片宽度
        
    def add_data(self, data):
        """添加统计数据"""
        if self.current_row == 1:
            headers = list(data[0].keys())
            self.ws.append(headers)
            self.current_row += 1
            
        for row in data:
            self.ws.append(list(row.values()))
            self.current_row += 1
            
    def _resize_image(self, img_path):
        """调整图片尺寸并返回临时文件路径"""
        img = Image.open(img_path)
        w, h = img.size
        if w > self.max_width:
            ratio = self.max_width / w
            new_size = (self.max_width, int(h * ratio))
            img = img.resize(new_size, Image.Resampling.LANCZOS)
            
        temp_buffer = BytesIO()
        img.save(temp_buffer, format='PNG')
        temp_buffer.seek(0)
        return temp_buffer
    
    def insert_images(self, folder_path, samples=10):
        """插入随机抽样图片"""
        # 创建统一的工作表
        sheet = self.wb.create_sheet(title="样本图片")
        
        # 收集所有类别数据
        categories = []
        for category_dir in os.listdir(folder_path):
            category_path = os.path.join(folder_path, category_dir)
            if os.path.isdir(category_path):
                all_images = glob.glob(os.path.join(category_path, '*.[pj][np]g'))
                selected = random.sample(all_images, min(samples, len(all_images)))
                categories.append((category_dir, selected))

        # 每行显示4个类别
        row_height = 0
        for idx, (cat_name, imgs) in enumerate(categories):
            # 计算行列位置
            col = (idx % 4) * 3 + 1  # 每3列一组（图片+文件名）
            group_row = (idx // 4) * (samples * 2 + 2)  # 每个类别组间隔2行
            
            # 写入类别标题
            cell = sheet.cell(group_row+1, col, f"{cat_name[:20]}...")
            sheet.merge_cells(start_row=group_row+1, start_column=col, 
                            end_row=group_row+1, end_column=col+1)
            sheet.cell(group_row+1, col, f"{cat_name[:20]}...")

            # 插入图片和文件名
            for img_idx, img_path in enumerate(imgs[:samples], 1):
                try:
                    img_buffer = self._resize_image(img_path)
                    img = xlImage(img_buffer)
                    
                    # 计算行位置
                    img_row = group_row + img_idx * 2 + 1
                    # 设置图片位置
                    sheet.add_image(img, f"{chr(64+col)}{img_row}")
                    # 写入文件名
                    sheet.cell(img_row+1, col, os.path.basename(img_path))
                    
                    # 记录最大行高
                    row_height = max(row_height, img.height * 0.75)
                    sheet.row_dimensions[img_row].height = row_height
                
                except Exception as e:
                    print(f"插入图片 {img_path} 失败: {str(e)}")

        # 调整列宽
        for col in sheet.columns:
            sheet.column_dimensions[col[0].column_letter].width = 30

    def adjust_columns(self):
        """调整列宽和格式"""
        # 原列宽调整逻辑
        for col in self.ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            adjusted_width = (max_length + 2)
            self.ws.column_dimensions[col[0].column_letter].width = adjusted_width
        
        # 类别列加粗
        bold_font = Font(bold=True)
        for row in self.ws.iter_rows(min_row=2, min_col=2, max_col=2):
            for cell in row:
                cell.font = bold_font

    def save(self):
        """保存文件并清理临时文件"""
        if os.path.exists(self.file_path):
            os.remove(self.file_path)
        self.wb.save(self.file_path)