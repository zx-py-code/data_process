import sys
from PyQt5.QtWidgets import *
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QFont, QPalette, QColor
import os
import shutil
import glob
from PIL import Image
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import Font
from PyQt5.QtCore import QDateTime

import os
import glob
import cv2
import numpy as np
from tqdm import tqdm
from openpyxl.drawing.image import Image as xlImage
import random
from io import BytesIO
# 新增Excel管理类
class ExcelManager:
    def __init__(self, file_path):
        # 新增文件存在检查逻辑
        if not os.path.exists(file_path):
            # 创建新工作簿并保存到指定路径
            self.wb = Workbook()
            self.wb.save(file_path)
        else:
            # 加载已有工作簿
            self.wb = load_workbook(file_path)
        self.ws = self.wb.active
        self.file_path = file_path
        self.current_row = 1
        self.max_width = 1024  # 最大图片宽度
        
    def add_data(self, data):
        """添加统计数据"""
        if self.current_row == 1:
            headers = list(data[0].keys())
            self.ws.append(headers)
            self.current_row += 1
            
        for row in data:
            self.ws.append(list(row.values()))
            self.current_row += 1
            
    def _resize_image(self, img_path):
        """调整图片尺寸并返回临时文件路径"""
        img = Image.open(img_path)
        w, h = img.size
        if w > self.max_width:
            ratio = self.max_width / w
            new_size = (self.max_width, int(h * ratio))
            img = img.resize(new_size, Image.Resampling.LANCZOS)
            
        temp_buffer = BytesIO()
        img.save(temp_buffer, format='PNG')
        temp_buffer.seek(0)
        return temp_buffer

    def insert_images(self, folder_path, samples=10):
        """插入随机抽样图片"""
        # 使用文件夹名称作为工作表名称
        sheet_name = os.path.basename(os.path.normpath(folder_path))
        sheet = self.wb.create_sheet(title=sheet_name[:30])  # 工作表名称最长31字符
        categories = []
        # 收集所有类别数据
        for category_dir in os.listdir(folder_path):
            category_path = os.path.join(folder_path, category_dir)
            if os.path.isdir(category_path):
                all_images = glob.glob(os.path.join(category_path, '*.bmp')) + glob.glob(os.path.join(category_path, '*.png'))
                selected = random.sample(all_images, min(samples, len(all_images)))
                categories.append((category_dir, selected))

        # 每行显示7个类别
        row_height = 0
        for idx, (cat_name, imgs) in enumerate(categories):
            # 计算行列位置
            col = (idx % 7) * 2 + 1  # 每3列一组（图片+文件名）
            group_row = (idx // 7) * (samples * 2 + 2)  # 每个类别组间隔2行
            cell = sheet.cell(group_row+1, col, f"{cat_name[:20]}")
            cell.fill = PatternFill(start_color="FFFF00", fill_type="solid")  # 仅当前单元格
            cell.font = Font(bold=True)  # 仅当前单元格

            # 插入图片和文件名
            for img_idx, img_path in enumerate(imgs[:samples], 1):
                try:
                    img_buffer = self._resize_image(img_path)
                    img = xlImage(img_buffer)
                    
                    # 计算行位置
                    img_row = group_row + img_idx * 2 + 1
                    # 设置图片位置
                    sheet.add_image(img, f"{chr(64+col)}{img_row}")
                    # 写入文件名
                    sheet.cell(img_row+1, col, os.path.basename(img_path))
                    
                    # 记录最大行高
                    row_height = max(row_height, img.height * 0.75)
                    sheet.row_dimensions[img_row].height = row_height
                
                except Exception as e:
                    print(f"插入图片 {img_path} 失败: {str(e)}")

        # 调整列宽
        for col in sheet.columns:
            sheet.column_dimensions[col[0].column_letter].width = 30

    def adjust_columns(self):
        """调整列宽和格式"""
        # 原列宽调整逻辑
        for col in self.ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            adjusted_width = (max_length + 2)
            self.ws.column_dimensions[col[0].column_letter].width = adjusted_width
        
        # 类别列加粗
        bold_font = Font(bold=True)
        for row in self.ws.iter_rows(min_row=2, min_col=2, max_col=2):
            for cell in row:
                cell.font = bold_font

    def save(self):
        """保存文件并清理临时文件"""
        if os.path.exists(self.file_path):
            os.remove(self.file_path)
        self.wb.save(self.file_path)
def cv_imread(file_path, type=0):
    #imdedcode读取的是RGB图像
    cv_img = cv2.imdecode(np.fromfile(file_path, dtype=np.uint8), -1)
    return cv_img


def cv_imwrite(file_path, img):
    cv2.imencode('.bmp', img)[1].tofile(file_path)

class ImageMaskVisualizer:
    def __init__(self, path, image_extensions=("png", "bmp", "jpeg"), mask_extension="png"):
        self.path = path
        self.image_folder = os.path.join(path, "JPEGImages")
        self.annotation_folder = os.path.join(path, "annotations")
        self.output_folder = os.path.join(path, "vis_img")
        self.cropped_folder = os.path.join(path, "crop_img")
        self.cropped_images_folder = os.path.join(self.cropped_folder, "JPEGImages")
        self.cropped_annotations_folder = os.path.join(self.cropped_folder, "Annotations")
        # Supported image extensions
        self.image_extensions = image_extensions
        self.mask_extension = mask_extension

    def visualize_masks(self):
        """Draw masks on images and save the results."""
        os.makedirs(self.output_folder, exist_ok=True)
        image_paths = []
        for ext in self.image_extensions:
            image_paths.extend(glob.glob(os.path.join(self.image_folder, f"*.{ext}")))

        for img_path in tqdm(image_paths):
            mask_path = img_path.replace("JPEGImages", "annotations").replace(os.path.splitext(img_path)[1], f".{self.mask_extension}")
            img = cv_imread(img_path, cv2.IMREAD_COLOR)
            mask = cv_imread(mask_path, cv2.IMREAD_GRAYSCALE)
            if len(mask.shape)== 3:
                mask = cv2.cvtColor(mask, cv2.COLOR_BGR2GRAY)
            else:
                print(mask.shape)
            if mask is None:
                print(f"Warning: Mask not found for {mask_path}. Skipping.")
                continue

            ret, thresh = cv2.threshold(mask, 0, 255, cv2.THRESH_BINARY)
            contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)

            img_with_contours = cv2.drawContours(img.copy(), contours, -1, (0, 0, 255), 2)

            output_path = os.path.join(self.output_folder, os.path.basename(img_path).replace(os.path.splitext(img_path)[1], "_valid.jpg"))
            cv_imwrite(output_path, img_with_contours)

        print(f"Processed images saved in: {self.output_folder}")


    def crop_images_from_masks(self, save_path = None,patch_size=1500, ratio=0.0001):
        # if save_path is None:
        #     save_path = os.path.join(self.path, 'crop_1500')
        #     os.makedirs(save_path, exist_ok=True)
        os.makedirs(self.cropped_images_folder, exist_ok=True)
        os.makedirs(self.cropped_annotations_folder, exist_ok=True)
        img_dir =self.image_folder
        label_dir = self.annotation_folder
        save_img_dir = self.cropped_images_folder
        save_label_dir = self.cropped_annotations_folder

        dir_list = os.listdir(label_dir)
        for file in tqdm(dir_list):
            # print(file)
            anno_path = os.path.join(label_dir, file)
            # anno = cv2.imread(anno_path, cv2.IMREAD_GRAYSCALE)
            anno = cv_imread(anno_path)
            img_path = os.path.join(img_dir, file.split('.png')[0]+'.bmp')
            # img = cv2.imread(img_path)  #
            img = cv_imread(img_path)  #
            H, W = anno.shape
            for h in range(0, H, patch_size):
                if h+patch_size > H:
                    h = H-patch_size

                for w in range(0, W, patch_size):
                    if w+patch_size > W:
                        w = W-patch_size

                    patch_anno = anno[h:h+patch_size, w:w+patch_size]
                    patch_anno = patch_anno * 1 # #
                    object_ratio = len(np.where(patch_anno > 0)[0]) / (patch_size*patch_size)
                    if object_ratio > ratio:
                        if len(img.shape) == 2:  # 检查图像是否为灰度图（二维数组）
                            patch_img = img[h:h + patch_size, w:w + patch_size]
                        else:  # 彩色图像（三维数组）
                            patch_img = img[h:h + patch_size, w:w + patch_size, :]
                        # patch_img = img[h:h + patch_size, w:w + patch_size, :]
                        print(patch_img.shape)
                        cv_imwrite(os.path.join(save_label_dir, file.split('.png')[0]+'_{}_{}.png').format(h, w), patch_anno)  ##
                        cv_imwrite(os.path.join(save_img_dir, file.split('.png')[0] + '_{}_{}.bmp').format(h, w), patch_img)

    # def _get_cropped_image(self, img, crop_x, crop_y, crop_size, is_mask=False):
    #     """Helper method to get a cropped image with black padding if necessary."""
    #     if is_mask:
    #         crop_img = np.zeros((crop_size[1], crop_size[0]), dtype=np.uint8)  # Black image for padding (1 channel for mask)
    #     else:
    #         crop_img = np.zeros((crop_size[1], crop_size[0], 3), dtype=np.uint8)  # Black image for padding (3 channels for color)

    #     img_height, img_width = img.shape[:2]

    #     # Define source coordinates with bounds checking
    #     src_x = max(crop_x, 0)
    #     src_y = max(crop_y, 0)
    #     src_w = min(crop_size[0], img_width - src_x)
    #     src_h = min(crop_size[1], img_height - src_y)

    #     # Copy the image part to the cropped image
    #     if is_mask:
    #         crop_img[max(-crop_y, 0):max(-crop_y, 0) + src_h,
    #                 max(-crop_x, 0):max(-crop_x, 0) + src_w] = img[src_y:src_y + src_h, src_x:src_x + src_w]
    #     else:
    #         crop_img[max(-crop_y, 0):max(-crop_y, 0) + src_h,
    #                 max(-crop_x, 0):max(-crop_x, 0) + src_w] = img[src_y:src_y + src_h, src_x:src_x + src_w]

    #     # If it's a mask, convert it to a 3-channel image for saving
    #     if is_mask:
    #         crop_img = cv2.cvtColor(crop_img, cv2.COLOR_GRAY2BGR)

    #     return crop_img

# 将原有功能封装到工作线程
class OrganizeWorker(QThread):
    progress = pyqtSignal(str)
    finished = pyqtSignal()
    error = pyqtSignal(str)

    def __init__(self, params):
        super().__init__()
        self.params = params

    

    def run(self):
        try:
            # 这里调用原始organize_folder函数，需稍作修改以支持进度信号
            self.organize_folder(self.params, self.progress.emit)
            self.finished.emit()
        except Exception as e:
            self.error.emit(str(e))
        
    def organize_folder(self,params,progress_callback=True):


        
        # 检查train、val和test文件夹是否存在
        train_path = os.path.join(params['folder_path'], 'train')
        # 检查train、val和test文件夹是否存在
        train_path = os.path.join(params['folder_path'], 'train')
        val_path = os.path.join(params['folder_path'], 'val')
        test_path = os.path.join(params['folder_path'], 'test')
        def report_progress(message):
            if progress_callback:
                progress_callback(message)

        if params['task_type'] == 'segmentation':
            # 处理分割任务
            for path in [train_path, val_path]:
                if os.path.exists(path):
                    # 创建必要的子文件夹
                    jpeg_path = os.path.join(path, 'JPEGImages')
                    annotations_path = os.path.join(path, 'annotations')
                    json_path = os.path.join(path, 'Json')
                    other_path = os.path.join(path, 'other')  # 创建用于存放其他类型文件的文件夹

                    for sub_path in [jpeg_path, annotations_path, json_path, other_path]:
                        print(sub_path)
                        os.makedirs(sub_path, exist_ok=True)

                    # 遍历指定文件夹下的所有文件
                    for root, _, files in os.walk(path):
                        for file in files:
                            file_path = os.path.join(root, file)
                            if file.endswith('.bmp'):
                                shutil.move(file_path, os.path.join(jpeg_path, file))
                            elif file.endswith('.png'):
                                shutil.move(file_path, os.path.join(annotations_path, file))
                            elif file.endswith('.json'):
                                shutil.move(file_path, os.path.join(json_path, file))
                            else:
                                shutil.move(file_path, os.path.join(other_path, file))
                    # 检查每个文件名是否存在三个版本
                    # 获取所有文件名（不带扩展名）
                    all_files = (
                        {os.path.splitext(f)[0] for f in os.listdir(jpeg_path) if f.endswith('.bmp')} |
                        {os.path.splitext(f)[0] for f in os.listdir(annotations_path) if f.endswith('.png')} |
                        {os.path.splitext(f)[0] for f in os.listdir(json_path) if f.endswith('.json')}
                    )
                    print(all_files)
                    for base_name in all_files:
                        expected_files = [
                            os.path.join(jpeg_path, f"{base_name}.bmp"),
                            os.path.join(annotations_path, f"{base_name}.png"),
                            os.path.join(json_path, f"{base_name}.json")
                        ]

                        # 统计存在的文件数量
                        existing_files = [f for f in expected_files if os.path.exists(f)]
                        print(existing_files)
                        # 如果存在但数量不足3个，移动到other
                        if 0 < len(existing_files) < 2:
                            print(existing_files)
                            for file_path in existing_files:
                                shutil.move(file_path, os.path.join(other_path, os.path.basename(file_path)))
                    # 删除空的other文件夹
                    if os.path.exists(other_path) and not os.listdir(other_path):
                        os.rmdir(other_path)
            if params['do_visualization']:
                print("可视化")
                for path in [train_path, val_path]:
                # 可视化
                    visualizer = ImageMaskVisualizer(path)
                    visualizer.visualize_masks()
            if params['do_crop']:
                print("切图")
                for path in [train_path, val_path]:
                # 切图
                    crop = ImageMaskVisualizer(path)
                    crop.crop_images_from_masks()
            if params['do_statistics']:
                # 统计文件数量并检查同名文件
                stats_data = []
                for path in [train_path, val_path]:
                    jpeg_path = os.path.join(path, 'JPEGImages')
                    annotations_path = os.path.join(path, 'annotations')
                    json_path = os.path.join(path, 'Json')

                    bmp_files = glob.glob(os.path.join(jpeg_path, '*.bmp'))
                    png_files = glob.glob(os.path.join(annotations_path, '*.png'))
                    json_files = glob.glob(os.path.join(json_path, '*.json'))

                    bmp_count = len(bmp_files)
                    png_count = len(png_files)
                    json_count = len(json_files)

                    stats_data.append({
                        'Folder': path,
                        'Number of .bmp files': bmp_count,
                        'Number of .png files': png_count,
                        'Number of .json files': json_count
                    })

                    # 检查同名文件
                    bmp_names = {os.path.splitext(os.path.basename(f))[0] for f in bmp_files}
                    png_names = {os.path.splitext(os.path.basename(f))[0] for f in png_files}
                    json_names = {os.path.splitext(os.path.basename(f))[0] for f in json_files}

                    all_names = bmp_names.union(png_names).union(json_names)
                    for name in all_names:
                        missing_info = {'Folder': path}
                        if name not in bmp_names:
                            missing_info['Missing file'] = f'{name}.bmp'
                        if name not in png_names:
                            missing_info['Missing file'] = f'{name}.png'
                        if name not in json_names:
                            missing_info['Missing file'] = f'{name}.json'
                        if 'Missing file' in missing_info:
                            stats_data.append(missing_info)

                stats_df = pd.DataFrame(stats_data)
                stats_file = os.path.join(params['folder_path'], 'statistics.xlsx')
                # 检查文件是否存在，如果存在则删除
                if os.path.exists(stats_file):
                    try:
                        os.remove(stats_file)
                    except Exception as e:
                        print(f"删除文件 {stats_file} 时出错: {e}")
                stats_df.to_excel(stats_file, index=False)

        elif params['task_type'] == 'classification':
            # 在关键步骤添加汇报，例如：
            report_progress("开始处理分类任务...")
            # 处理分类任务
            all_counts = 0
            all_image_sizes = set()
            stats_data = []
            other_path = os.path.join(params['folder_path'], 'other')
            os.makedirs(other_path, exist_ok=True)

            for path in [train_path, val_path, test_path]:
                if os.path.exists(path):
                    category_folders = [d for d in os.listdir(path) if os.path.isdir(os.path.join(path, d))]
                    for category in category_folders:
                        category_path = os.path.join(path, category)
                        for item in os.listdir(category_path):
                            item_path = os.path.join(category_path, item)
                            if os.path.isdir(item_path) or not item.lower().endswith(('.bmp', '.png')):
                                shutil.move(item_path, os.path.join(other_path, item))

                        image_files = glob.glob(os.path.join(category_path, '*.bmp')) + glob.glob(os.path.join(category_path, '*.png'))
                        image_count = len(image_files)
                        all_counts += image_count
                        formats = ['.bmp', '.png']
                        if params['convert_format']:
                            if params['convert_format'] == 'png_to_bmp':
                                for img_file in glob.glob(os.path.join(category_path, '*.png')):
                                    try:
                                        img = Image.open(img_file)
                                        new_img_path = os.path.splitext(img_file)[0] + '.bmp'
                                        img.save(new_img_path, 'BMP')
                                        os.remove(img_file)
                                    except Exception as e:
                                        print(f"转换 {img_file} 时出错: {e}")
                                formats = ['.bmp']
                            elif params['convert_format'] == 'bmp_to_png':
                                for img_file in glob.glob(os.path.join(category_path, '*.bmp')):
                                    try:
                                        img = Image.open(img_file)
                                        new_img_path = os.path.splitext(img_file)[0] + '.png'
                                        img.save(new_img_path, 'PNG')
                                        os.remove(img_file)
                                    except Exception as e:
                                        print(f"转换 {img_file} 时出错: {e}")
                                formats = ['.png']
                            report_progress("完成图像格式转换")

                        # 统计所有图像格式的数量

                        format_counts = {}

                        # 统计所有图像格式的数量
                        
                        format_counts = {}
                        for fmt in formats:
                            format_files = glob.glob(os.path.join(category_path, f'*{fmt}'))
                            format_counts[fmt[1:]] = len(format_files)

                        # 找到数量最少的图像格式
                        min_format = ''
                        min_count = 0
                        if format_counts:
                            min_format = min(format_counts, key=format_counts.get)
                            min_count = format_counts[min_format]

                        category_image_sizes = set()
                        image_files = glob.glob(os.path.join(category_path, '*.bmp')) + glob.glob(os.path.join(category_path, '*.png'))
                        for img_file in image_files:
                            try:
                                with Image.open(img_file) as img:
                                    width, height = img.size
                                    category_image_sizes.add((width, height))
                                    all_image_sizes.add((width, height))
                            except Exception as e:
                                print(f"Error opening {img_file}: {e}")

                        size_str = ", ".join([f"{w}x{h}" for w, h in category_image_sizes])
                                            # 执行图像格式转换
                        # 执行图像重命名功能
                        if params['do_rename']:
                            if 'rename_rule' in params:
                                rename_template = params['rename_rule']
                                for index, img_file in enumerate(glob.glob(os.path.join(category_path, '*.*'))):
                                    file_ext = os.path.splitext(img_file)[1]
                                    base_new_name = rename_template.format(
                                        category=category,
                                        index=index,
                                    )
                                    new_name = f"{base_new_name}{file_ext}"
                                    new_path = os.path.join(category_path, new_name)
                                    counter = 1
                                    # 检查新文件名是否已存在
                                    while os.path.exists(new_path):
                                        new_name = f"{base_new_name}_{counter}{file_ext}"
                                        new_path = os.path.join(category_path, new_name)
                                        counter += 1
                                    # print(f"Renaming {img_file} to {new_name}")
                                    os.rename(img_file, new_path)
                            else:
                                # 原有的重命名规则
                                for index, img_file in enumerate(glob.glob(os.path.join(category_path, '*.*'))):
                                    file_ext = os.path.splitext(img_file)[1]
                                    base_new_name = f"{category}_{index}"
                                    new_name = f"{base_new_name}{file_ext}"
                                    new_path = os.path.join(category_path, new_name)
                                    counter = 1
                                    # 检查新文件名是否已存在
                                    while os.path.exists(new_path):
                                        new_name = f"{base_new_name}_{counter}{file_ext}"
                                        new_path = os.path.join(category_path, new_name)
                                        counter += 1
                                    os.rename(img_file, new_path)
                            # report_progress("完成图像重命名")
                        stats_data.append({
                            'Folder': path,
                            'Category': category,
                            'Image Count': image_count,
                            'Min format': min_format,
                            'Min Count': min_count,
                            **{f'{fmt} format Count': count for fmt, count in format_counts.items()},
                            'Image sizes in category': size_str
                        })
            dataset_size_str = ", ".join([f"{w}x{h}" for w, h in all_image_sizes])

            if params['do_statistics']:
                stats_data.append({
                    'Folder': 'Dataset',
                    'Category': '',
                    'Image Count': all_counts,
                    'Min format': '',
                    'Min Count': '',
                    'Image sizes in category': dataset_size_str
                })
                stats_file = os.path.join(params['folder_path'], 'statistics.xlsx')
                # stats_df = pd.DataFrame(stats_data)
                # 检查文件是否存在，如果存在则删除
                if os.path.exists(stats_file):
                    try:
                        os.remove(stats_file)
                    except Exception as e:
                        print(f"删除文件 {stats_file} 时出错: {e}")
                # stats_df.to_excel(stats_file, index=False)
                excel_mgr = ExcelManager(stats_file)
                # stats_df = pd.DataFrame(stats_data)
                # stats_file = os.path.join(params['folder_path'], 'statistics.xlsx')
                # 添加统计数据
                excel_mgr.add_data(stats_data)
                
                # 如果是分类任务添加图片
                if params['task_type'] == 'classification':
                    for path in [train_path, val_path, test_path]:
                        if os.path.exists(path):
                            excel_mgr.insert_images(path)
                
                # 调整格式并保存
                excel_mgr.adjust_columns()
                excel_mgr.save()
                # # 检查文件是否存在，如果存在则删除
                # if os.path.exists(stats_file):
                #     try:
                #         os.remove(stats_file)
                #     except Exception as e:
                #         print(f"删除文件 {stats_file} 时出错: {e}")
                # stats_df.to_excel(stats_file, index=False)

                # # 加载生成的 Excel 文件并调整格式
                # wb = load_workbook(stats_file)
                # ws = wb.active

                # # 调整列宽以适应内容
                # for column in ws.columns:
                #     max_length = 0
                #     column_letter = column[0].column_letter
                #     for cell in column:
                #         try:
                #             if len(str(cell.value)) > max_length:
                #                 max_length = len(str(cell.value))
                #         except:
                #             pass
                #     adjusted_width = (max_length + 2)
                #     ws.column_dimensions[column_letter].width = adjusted_width
                # 类别列（假设是第二列）加粗处理
                # bold_font = Font(bold=True)
                # for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
                #     for cell in row:
                #         cell.font = bold_font
                # 保存修改后的 Excel 文件
                # wb.save(stats_file)
        else:
            print(f"不支持的任务类型: {params['task_type']}")